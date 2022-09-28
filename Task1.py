#Import pandas lib as pd
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import PatternFill
#from itertools import chain


#Function to read files

def reading_files():

	name_file1 = '1MC03-SCJ-IM-REP-SS01_SL03-242400_P07_Data.xlsx'
	name_file2 = '1MC03-SCJ-IM-REP-SS01_SL03-242400_P08_Data.xlsx'

	all_dfs1 = pd.read_excel(name_file1, sheet_name = None)
	all_dfs2 = pd.read_excel(name_file2, sheet_name = None)

	print(all_dfs1)

	#Converting into one tab 

	#all_dfs1.keys()
	#all_dfs2.keys()

	df_1 = pd.concat(all_dfs1, ignore_index = True)
	df_2= pd.concat(all_dfs2, ignore_index = True)


	# Adding 'row_from_file' and 'Identify GUI'


	df_1['Row_from_file'] = '1MC03-SCJ-IM-REP-SS01_SL03-242400_P07_Data.xlsx'
	df_2['Row_from_file' ]= '1MC03-SCJ-IM-REP-SS01_SL03-242400_P08_Data.xlsx'

	df_1.insert(1, 'Identify GUID','')
	df_1['Identify GUID'] = 'old'
	df_2.insert(1, 'Identify GUID','')
	df_2['Identify GUID'] = 'new'


	# Combining the files (new and old) into one


	excel_merged = df_1.append(df_2)
	writer = pd.ExcelWriter('Comparison.xlsx', engine='xlsxwriter')
	excel_merged.to_excel(writer)

	#Writing the fil1
	file1 = pd.ExcelWriter(name_file1, engine='xlsxwriter')
	df_1.to_excel(file1)

	file2 = pd.ExcelWriter(name_file2, engine='xlsxwriter')
	df_2.to_excel(file2)

	file1.save()
	file2.save()
	writer.save()


# Comparing GUID column and using tags to identify same/new/old

def same_GUID():

	comparison_file = 'Comparison.xlsx'
	writer_1 = pd.ExcelWriter(comparison_file, engine='openpyxl')
	info = pd.read_excel(comparison_file, header = 0)
	info.sort_values(by=['GUID'], inplace = True)
	info.loc[info['GUID'].duplicated(keep= False), 'Identify GUID' ] = 'same'
	info.drop('Unnamed: 0', axis=1, inplace=True)
	#info.style.apply(highlight_rows, axis=1)
	info.to_excel(writer_1)
	writer_1.save()


def comparison_sameGUID():
	c=0
	out_pos = {}
	out_values = {}

	GUID = {}
	GUID_pos = {}

	wb1 = openpyxl.load_workbook('Comparison.xlsx') #Loding the file openpyxl
	fill_cell = PatternFill(patternType= 'solid', 
                            fgColor='00CCCCFF') #Creating the pattern for filling the cells


	wb = wb1.active # Get the unique sheet on the workbook
	#Create a dictionary of column names 
	ColNames = {}
	#Cont variable to move between the columns inside the workbook
	Current = 0
	cont_1 = 0
	#Saving the columns name in a dictionary
	for COL in wb.iter_cols(1, wb.max_column):
		ColNames[COL[0].value] = Current
		Current += 1
	#Colouring rows when the 'Identify GUID' column value is equal to same 
	for row_cells in wb.iter_rows(1, wb.max_row):
		GUID[cont_1] = row_cells[ColNames['GUID']].value
		GUID_pos[cont_1] = row_cells[ColNames['GUID']]
		if (row_cells[ColNames['Identify GUID']].value) == 'same':
			for cell in row_cells:
				cell.fill = fill_cell
				#out_pos[cont_1]= cell
				#out_values[cont_1] = cell.value
				#cont_1+=1
		cont_1+=1


	print(GUID_pos)
	#out_pos = out_pos.values()

	repeated_values(out_values,out_pos)


	#print(out_pos.values())
	#print(out_values)
	#print(out_values)

	#print(len(out_pos))

	wb1.save('Task_1.xlsx')
	wb1.close()



def repeated_values(b,c):

	fill_cell2 = PatternFill(patternType= 'solid', 
                            fgColor='FFC7CE')
	rev_multidic = {}
	valor = []
	for key, value in b.items():
		rev_multidic.setdefault(value, set()).add(key)





	#keys_repeated = [key for key, values in rev_multidic.items() if len(values) > 1 ]
	values_repeated = [values for key , values in rev_multidic.items() if len(values) == 1]

	for x in range(len(values_repeated)):
		valor.append(list(values_repeated[x]))

	for j in range (len(valor)):
		cell = c[valor[j][0]]
		cell.fill = fill_cell2
 
#a = reading_files()
#b = same_GUID()
d = comparison_sameGUID()